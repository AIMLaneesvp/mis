import os
import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# ===================================
# ⚙️ Configuration
# ===================================
DATA_FOLDER = r"C:\MIS_Files"  # 👉 Change this to your desired folder
# DATA_FOLDER = "/home/ubuntu/mis_files"  # for Linux servers
os.makedirs(DATA_FOLDER, exist_ok=True)
EXCEL_PATH = os.path.join(DATA_FOLDER, "Book1.xlsx")

APP_USERNAME = "admin"
APP_PASSWORD = "admin123"

# ===================================
# 🔒 Authentication (Stable Auto-Redirect)
# ===================================
def password_gate():
    """Secure login system with username + password and safe rerun after login."""
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if "username" not in st.session_state:
        st.session_state.username = None

    # If already logged in — show logout
    if st.session_state.authenticated:
        with st.sidebar:
            st.success(f"👋 Logged in as {st.session_state.username}")
            if st.button("Logout"):
                st.session_state.authenticated = False
                st.session_state.username = None
                st.experimental_rerun()
        return True

    # --- Login UI ---
    st.title("🔐 Secure MIS System Login")
    st.markdown("Please enter your credentials to access the MIS system.")

    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    login_btn = st.button("Login")

    if login_btn:
        if username == APP_USERNAME and password == APP_PASSWORD:
            st.session_state.authenticated = True
            st.session_state.username = username
            st.success("✅ Access Granted! Redirecting...")
            st.experimental_rerun()   # immediately load MIS system
        else:
            st.error("❌ Invalid credentials. Try again.")

    st.stop()

# Run authentication
if not password_gate():
    st.stop()

# ===================================
# 🔧 Helper: Append Data to Excel
# ===================================
def append_df_to_excel(filename, df, sheet_name):
    """Append a DataFrame to an existing Excel sheet or create new one."""
    if not os.path.exists(filename):
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        return
    book = load_workbook(filename)
    if sheet_name not in book.sheetnames:
        with pd.ExcelWriter(filename, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
    else:
        start_row = book[sheet_name].max_row
        with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=start_row)

# ===================================
# 💹 Summary Function
# ===================================
def calculate_summary(purchase_df, sales_df):
    """Convert numeric columns safely and compute totals."""
    for df in [purchase_df, sales_df]:
        for col in ["INR", "BHD", "Quantity", "Item Rate (BHD)"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    total_purchase = purchase_df["INR"].sum() if "INR" in purchase_df.columns else 0
    total_sales = sales_df["INR"].sum() if "INR" in sales_df.columns else 0
    gross_profit = total_sales - total_purchase
    return {
        "Total Purchase (INR)": total_purchase,
        "Total Sales (INR)": total_sales,
        "Gross Profit (INR)": gross_profit,
    }

# ===================================
# 🧱 Streamlit App Layout
# ===================================
st.set_page_config(page_title="MIS System", layout="wide")
st.title("📊 MIS System - Purchase & Sales Entry")
st.caption(f"📁 Data is stored locally at: `{EXCEL_PATH}`")

# Initialize session state lists
if "purchase_entries" not in st.session_state:
    st.session_state.purchase_entries = []
if "sales_entries" not in st.session_state:
    st.session_state.sales_entries = []

# ==============================
# 🛒 PURCHASE SECTION
# ==============================
st.subheader("🧾 Purchase Entry")

with st.form("purchase_form"):
    c1, c2, c3, c4, c5, c6 = st.columns([1.2, 1.5, 1, 1, 1, 1])
    with c1:
        p_date = st.date_input("Date", key="p_date")
    with c2:
        p_vendor = st.text_input("Vendor Name", key="p_vendor")
    with c3:
        p_item_rate = st.number_input("Item Rate (BHD)", min_value=0.0, format="%.3f", key="p_rate")
    with c4:
        p_quantity = st.number_input("Quantity", min_value=0.0, format="%.2f", key="p_qty")
    with c5:
        p_inr = p_quantity * 1000
        st.number_input("INR Value (auto)", value=p_inr, disabled=True, format="%.2f")
    with c6:
        p_bhd = p_item_rate * p_quantity
        st.number_input("BHD (auto)", value=p_bhd, disabled=True, format="%.3f")

    submitted_purchase = st.form_submit_button("➕ Add Purchase Entry")

if submitted_purchase:
    if p_vendor and p_item_rate > 0 and p_quantity > 0:
        st.session_state.purchase_entries.append({
            "Date": p_date,
            "Vendor": p_vendor,
            "Item Rate (BHD)": p_item_rate,
            "Quantity": p_quantity,
            "INR": p_inr,
            "BHD": p_bhd
        })
        st.success("✅ Purchase entry added successfully!")
    else:
        st.warning("⚠️ Please fill all purchase fields correctly.")

if st.session_state.purchase_entries:
    df_pur = pd.DataFrame(st.session_state.purchase_entries)
    st.dataframe(df_pur)
    st.write(f"**Total INR:** ₹{df_pur['INR'].sum():,.2f} | **Total BHD:** {df_pur['BHD'].sum():,.3f}")

# ==============================
# 💰 SALES SECTION
# ==============================
st.subheader("💰 Sales Entry")

with st.form("sales_form"):
    s1, s2, s3, s4, s5, s6 = st.columns([1.2, 1.5, 1, 1, 1, 1])
    with s1:
        s_date = st.date_input("Date", key="s_date")
    with s2:
        s_customer = st.text_input("Customer Name", key="s_customer")
    with s3:
        s_item_rate = st.number_input("Item Rate (BHD)", min_value=0.0, format="%.3f", key="s_rate")
    with s4:
        s_quantity = st.number_input("Quantity", min_value=0.0, format="%.2f", key="s_qty")
    with s5:
        s_inr = s_quantity * 1000
        st.number_input("INR Value (auto)", value=s_inr, disabled=True, format="%.2f")
    with s6:
        s_bhd = s_item_rate * s_quantity
        st.number_input("BHD (auto)", value=s_bhd, disabled=True, format="%.3f")

    submitted_sales = st.form_submit_button("➕ Add Sales Entry")

if submitted_sales:
    if s_customer and s_item_rate > 0 and s_quantity > 0:
        st.session_state.sales_entries.append({
            "Date": s_date,
            "Customer": s_customer,
            "Item Rate (BHD)": s_item_rate,
            "Quantity": s_quantity,
            "INR": s_inr,
            "BHD": s_bhd
        })
        st.success("✅ Sales entry added successfully!")
    else:
        st.warning("⚠️ Please fill all sales fields correctly.")

if st.session_state.sales_entries:
    df_sales = pd.DataFrame(st.session_state.sales_entries)
    st.dataframe(df_sales)
    st.write(f"**Total INR:** ₹{df_sales['INR'].sum():,.2f} | **Total BHD:** {df_sales['BHD'].sum():,.3f}")

# ==============================
# 💾 SAVE TO EXCEL
# ==============================
st.markdown("---")
st.subheader("💾 Save Data")

save_clicked = st.button("💾 Save All to Excel (No Rerun Until Click)")

if save_clicked:
    if st.session_state.purchase_entries:
        append_df_to_excel(EXCEL_PATH, pd.DataFrame(st.session_state.purchase_entries), "Purchase")
    if st.session_state.sales_entries:
        append_df_to_excel(EXCEL_PATH, pd.DataFrame(st.session_state.sales_entries), "Sales")

    st.session_state.purchase_entries = []
    st.session_state.sales_entries = []
    st.success(f"✅ Data saved successfully to `{EXCEL_PATH}`!")
    st.experimental_rerun()  # refresh after save

# ==============================
# 📊 SUMMARY
# ==============================
if os.path.exists(EXCEL_PATH):
    try:
        xls = pd.ExcelFile(EXCEL_PATH)
        pur = pd.read_excel(xls, "Purchase") if "Purchase" in xls.sheet_names else pd.DataFrame()
        sal = pd.read_excel(xls, "Sales") if "Sales" in xls.sheet_names else pd.DataFrame()

        for df in [pur, sal]:
            for col in ["INR", "BHD", "Quantity", "Item Rate (BHD)"]:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        summary = calculate_summary(pur, sal)

        st.markdown("---")
        st.write("### 📈 Summary Overview")
        colA, colB, colC = st.columns(3)
        colA.metric("Total Purchase (INR)", f"{summary['Total Purchase (INR)']:.2f}")
        colB.metric("Total Sales (INR)", f"{summary['Total Sales (INR)']:.2f}")
        colC.metric("Gross Profit (INR)", f"{summary['Gross Profit (INR)']:.2f}")
    except Exception as e:
        st.warning(f"⚠️ Could not read summary: {e}")
else:
    st.info("ℹ️ Add and save entries to view summary.")
